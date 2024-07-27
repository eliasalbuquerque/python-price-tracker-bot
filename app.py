"""
title: 'Robô de Monitoramento Diário de Preço DestravaDev#3'
author: 'Elias Albuquerque'
version: 'Python 3.12.0'
created: '2024-07-25'
update: '2024-07-27'
"""


import logging.config
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.keys import Keys
from selenium.common.exceptions import *
from time import sleep
import datetime
import os
import sys
import openpyxl
import schedule
import keyboard


def logging_settings():
    """
    Configura o logging da aplicação utilizando o arquivo 'config.ini'.
    """

    logging.config.fileConfig('config.ini', disable_existing_loggers=False)
    logging.warning('Aplicação inicializada')

def driver_settings():
    """
    Configura o driver do Chrome e o WebDriverWait para interação com o site.
    Returns: driver, wait
    """

    logging.info('Iniciando configurações da aplicação...')
    try:
        options = Options()
        arguments = [
            '--block-new-web-contents',
            '--disable-notifications',
            '--no-default-browser-check',
            '--lang=pt-BR',
            # '--headless',
            '--window-position=36,68',
            '--window-size=1100,750',]

        for argument in arguments:
            options.add_argument(argument)

        options.add_experimental_option("excludeSwitches", ["enable-logging"])

        driver = webdriver.Chrome(options=options)
        wait = WebDriverWait(
            driver,
            15,
            poll_frequency=1,
            ignored_exceptions=[
                NoSuchElementException,
                ElementNotVisibleException,
                ElementNotSelectableException])

        return driver, wait

    except Exception as e:
        logging.error(f'Erro na configuração do driver: {e}')
        return None, None

def access_website(url, driver, wait):
    """
    Esta função acessa um site especificado usando um driver de navegador.
    Args: url (str), driver (webdriver), wait (int)
    Returns: driver (webdriver)
    """

    logging.info('Acessando o site (aguarde!)...')
    try:
        driver.get(url)
        sleep(10)
        driver.execute_script(f'document.body.style.zoom=".67"')
        sleep(5)

        return driver

    except TimeoutException as e:
        logging.error(f'Erro ao acessar o site {url}: {e}')
        return None
    except NoSuchElementException as e:
        logging.error(f'Erro ao acessar o site {url}: Elemento não encontrado: {e}')
        return None
    except WebDriverException as e:
        logging.error(f'Erro ao acessar o site {url}: {e}')
        return None

def extract_product_value(driver, xpath_element):
    """
    Extrai o valor do produto da página web usando JavaScript.
    Args: driver (webdriver)
    Returns: price_value (str)
    """

    logging.info('Extraindo o valor do produto...')
    try:
        price_value = driver.execute_script(
            "return document.querySelector(arguments[0]).getAttribute('content')", 
            xpath_element)

        return price_value

    except NoSuchElementException as e:
        logging.error(f'Erro ao acessar o valor do produto: {str(e).split("\n")[0]}')
        return None
    except WebDriverException as e:
        logging.error(f'Erro ao acessar o valor do produto: {str(e).split("\n")[0]}')
        return None
    finally:
        driver.quit()

def process_data(data):
    """
    Converte a string de dados para inteiro ou decimal.
    Args: data (str)
    Returns: data_value (int or float)
    """

    if "." in data or "," in data:
        data_value = float(data)
    else:
        data_value = int(data)

    return data_value

def create_spreadsheet(spreadsheet_name, sheet_name, columns):
    """
    Cria uma planilha Excel com colunas especificadas.
    Args: spreadsheet_name (str), sheet_name (str), columns (list)
    Returns: spreadsheet_name (str)
    """

    if os.path.exists(spreadsheet_name):
        return spreadsheet_name

    try:
        logging.info(f'Criando nova planilha "{spreadsheet_name}"...')

        workbook = openpyxl.Workbook()
        workbook.active.title = sheet_name
        sheet = workbook.active
        sheet.append(columns)

        # Aplicando estilo da planilha
        sheet.row_dimensions[1].height = 30
        for col in range(1, len(columns) + 1):
            sheet.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 20

        for row in sheet.iter_rows():
            for cell in row:
                cell.font = openpyxl.styles.Font(name='Calibri', size=13, bold=True)
                cell.alignment = openpyxl.styles.Alignment(
                    horizontal='center', vertical='center')
                cell.border = openpyxl.styles.Border(
                    top=openpyxl.styles.Side(style='thin'), 
                    bottom=openpyxl.styles.Side(style='thin'))  # Borda

        workbook.save(spreadsheet_name)
        return spreadsheet_name

    except Exception as e:
        logging.error(f'Erro ao criar a planilha: {e}')
        return None

def generating_data_for_spreadsheet(product_name, price, url_link):
    """
    Gera um dicionário com os dados do produto para inserir na planilha.
    Args: product_name (str), price (int or float), url_link (str)
    Returns: data_product (dict)
    """

    data_product = {
        'product': product_name,
        'date': datetime.datetime.now().strftime('%d/%m/%y %H:%M:%S'),
        'value': price,
        'link': url_link
    }
    return data_product

def insert_data_into_spreadsheet(spreadsheet_name, data_to_insert):
    """
    Insere os dados na planilha Excel.
    Args: spreadsheet_name (str), data_to_insert (dict)
    """

    logging.info(f'Inserindo dados na planilha...')

    try:
        workbook = openpyxl.load_workbook(spreadsheet_name)
        sheet = workbook.active

        row_num = sheet.max_row + 1
        for col_num, value in enumerate(data_to_insert.values(), start=1):
            if col_num == 4:
                sheet.cell(row=row_num, column=col_num).hyperlink = value
                sheet.cell(row=row_num, column=col_num).alignment = openpyxl.styles.Alignment(horizontal='center')
            else:
                sheet.cell(row=row_num, column=col_num).value = value
                sheet.cell(row=row_num, column=col_num).alignment = openpyxl.styles.Alignment(horizontal='center')

        workbook.save(spreadsheet_name)

    except Exception as e:
        logging.error(f'Erro ao inserir dados na planilha: {e}')

def schedule_application_execution(minutes=30):
    """
    Agendar a execução da aplicação a cada intervalo de minutos definido.
    Args: minutes (int): Intervalo de minutos para execução com valor padrão de 30 minutos.
    """
    try:
        logging.info(f'Agendando execução da aplicação para daqui à {minutes} minutos...')

        def run_application():
            try:
                logging.info('Executando o script agendado...')
                main()
            except Exception as e:
                logging.error(f'Erro durante a execução da aplicação: {e}')
            finally:
                schedule.every(minutes).minutes.do(run_application).tag('application')

        # Cancelar qualquer agendamento anterior
        schedule.clear('application')

        # Agendar a execução inicial
        schedule.every(minutes).minutes.do(run_application).tag('application')

        # Obter o próximo horário da próxima execução
        next_execution = schedule.next_run()
        next_time_execution = next_execution.strftime('%H:%M:%S')
        logging.info(f'A aplicação será executada às: {next_time_execution}')

        print('\t - Mantenha pressionada a tecla:')
        print('\t\t"ESC": Para interromper a execução')
        print('\t\t"P"  : Para abrir a planilha')
        
        while True:
            schedule.run_pending()

            # Abrir a planilha
            if keyboard.is_pressed('p'):
                os.system('powershell -Command "start \'.\\Registro de preços.xlsx\'"')

            # Interromper a execução
            if keyboard.is_pressed('esc'):
                schedule.clear('application')
                logging.info("Script interrompido pelo usuário")
                logging.info('Aplicação finalizada\n')
                sys.exit()

            sleep(1)

    except Exception as e:
        logging.error(f'Erro ao agendar execução da aplicação: {e}')

def main():
    # url = 'https://bit.ly/tabS9ultra_int_value' # int value
    url = 'https://bit.ly/tabS9ultra'  # float value
    xpath_element = 'meta[itemprop="price"]'
    spreadsheet_name = 'Registro de preços.xlsx'
    sheet_name = 'Produto'
    columns = ['Produto', 'Data Atual', 'Valor', 'Link Produto']
    product_name = 'Tab S9 Ultra'
    

    # 0. Configurações da aplicação
    logging_settings()
    driver, wait = driver_settings()

    # 1. Acessar o site
    access_website(url, driver, wait)

    if not driver:
        return

    # 2. Coletar o preço do produto
    price_value = extract_product_value(driver, xpath_element)

    if not price_value:
        return

    # 3. Tratar o dado para número inteiro ou decimal
    price = process_data(price_value)

    if not price:
        return

    # 4. Criar planilha com as colunas: Produto, Data Atual, Valor, Link Produto
    spreadsheet_name = create_spreadsheet(spreadsheet_name, sheet_name, columns)

    if not spreadsheet_name:
        return

    # 5. Inserir os dados na planilha e salvar
    data_product = generating_data_for_spreadsheet(product_name, price, url)
    insert_data_into_spreadsheet(spreadsheet_name, data_product)

    # 6. Agendar o script para que execute a cada 30 min.
    schedule_application_execution(2)

if __name__ == '__main__':
    main()
